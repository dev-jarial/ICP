import csv

import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import path, reverse
from django.utils.html import format_html

from .models import Company, CompanyUpload
from .views import upload_company_file

admin.site.site_header = "Partner Profile"
admin.site.site_title = "CoPartner Profile Admin Portal"
admin.site.index_title = "Welcome to Partner Profile Dashboard"


@admin.action(description="Export selected companies as CSV")
def export_as_csv(modeladmin, request, queryset):
    """Export selected company data as CSV"""
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = "attachment; filename=selected_companies.csv"

    writer = csv.writer(response)

    # Define headers
    headers = [
        field.name for field in Company._meta.fields
    ]  # Fetch all field names dynamically
    writer.writerow(headers)

    # Write data rows
    for company in queryset:
        row = [getattr(company, field) for field in headers]
        writer.writerow(row)

    return response


@admin.action(description="Export selected companies as Excel")
def export_as_excel(modeladmin, request, queryset):
    """Export selected company data as Excel"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Companies"

    # Define headers
    headers = [
        field.name for field in Company._meta.fields
    ]  # Fetch all field names dynamically
    sheet.append(headers)

    # Add data rows
    for company in queryset:
        row = []
        for field in headers:
            value = getattr(company, field)
            if isinstance(value, list) or isinstance(value, dict):
                value = str(value)  # Convert JSONField data to string
            row.append(value)
        sheet.append(row)

    # Prepare the response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = "attachment; filename=selected_companies.xlsx"

    workbook.save(response)
    return response


class CompanyAdmin(admin.ModelAdmin):
    list_display = ("edit_button", "name", "email", "download_links", "view_details")
    search_fields = ("name", "email")
    ordering = ("id",)
    actions = [export_as_csv, export_as_excel]

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path("export/csv/", self.export_all_as_csv, name="export_csv"),
            path("export/excel/", self.export_all_as_excel, name="export_excel"),
            path(
                "export/csv/<int:company_id>/",
                self.admin_site.admin_view(self.export_csv_single),
                name="export_csv_single",
            ),
            path(
                "export/excel/<int:company_id>/",
                self.admin_site.admin_view(self.export_excel_single),
                name="export_excel_single",
            ),
            path(
                "upload/",
                self.admin_site.admin_view(upload_company_file),
                name="upload_company_file",
            ),
            path(
                "<int:company_id>/view/",
                self.admin_site.admin_view(self.company_admin_view),
                name="company_admin_view",
            ),
        ]
        return custom_urls + urls

    def edit_button(self, obj):
        edit_url = reverse(
            "admin:company_company_change", args=[obj.id]
        )  # Generates Edit URL
        return format_html(
            '<a href="{}" class="button" title="Edit Company"><i class="fa fa-edit"></i> Edit</a>',
            edit_url,
        )

    edit_button.short_description = "Edit"  # Column title in Django Admin

    def view_details(self, obj):
        view_url = reverse(
            "admin:company_admin_view", args=[obj.id]
        )  # Ensure correct URL pattern
        return format_html(
            '<a href="{}" title="View Details"><i class="fa fa-eye"></i> View</a>',
            view_url,
        )

    view_details.short_description = "View"

    def company_admin_view(self, request, company_id):
        company = get_object_or_404(Company, id=company_id)
        return render(
            request, "admin/company_admin_view.html", {"company": company}
        )  # Updated template name

    def export_csv_single(self, request, company_id):
        """Export a single company's data as CSV"""
        company = get_object_or_404(Company, id=company_id)

        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f"attachment; filename={company.name}_data.csv"
        )

        writer = csv.writer(response)

        # Define headers
        headers = [field.name for field in Company._meta.fields]
        writer.writerow(headers)

        # Write single company's data
        row = [getattr(company, field) for field in headers]
        writer.writerow(row)

        return response

    def export_excel_single(self, request, company_id):
        """Export a single company's data as Excel"""
        company = get_object_or_404(Company, id=company_id)

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Company Data"

        # Define headers
        headers = [field.name for field in Company._meta.fields]
        sheet.append(headers)

        # Write single company's data
        row = []
        for field in headers:
            value = getattr(company, field)
            if isinstance(value, list) or isinstance(value, dict):
                value = str(value)  # Convert JSONField data to string
            row.append(value)
        sheet.append(row)

        # Prepare the response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            f"attachment; filename={company.name}_data.xlsx"
        )

        workbook.save(response)
        return response

    def export_all_as_csv(self, request):
        """Export all company data as CSV"""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=all_companies.csv"

        writer = csv.writer(response)

        headers = [
            field.name for field in Company._meta.fields
        ]  # Fetch all field names dynamically
        writer.writerow(headers)

        for company in Company.objects.all():
            row = [getattr(company, field) for field in headers]
            writer.writerow(row)

        return response

    def export_all_as_excel(self, request):
        """Export all company data as Excel"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Companies"

        headers = [
            field.name for field in Company._meta.fields
        ]  # Fetch all field names dynamically
        sheet.append(headers)

        for company in Company.objects.all():
            row = []
            for field in headers:
                value = getattr(company, field)
                if isinstance(value, list) or isinstance(value, dict):
                    value = str(value)  # Convert JSONField data to string
                row.append(value)
            sheet.append(row)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=all_companies.xlsx"

        workbook.save(response)
        return response

    def download_links(self, obj):
        """Adds download buttons in the admin panel for individual companies"""
        csv_url = reverse("admin:export_csv_single", args=[obj.id])  # Pass company ID
        excel_url = reverse(
            "admin:export_excel_single", args=[obj.id]
        )  # Pass company ID
        return format_html(
            f'<a href="{csv_url}">📥 Download CSV</a> | <a href="{excel_url}">📥 Download Excel</a>'
        )

    download_links.short_description = "Download Table"

    def upload_view(self, request):
        return redirect(reverse("upload_company_file"))

    def upload_files_link(self, obj):
        """Adds an upload button in the admin panel"""
        url = reverse("upload_company_file")
        return format_html(f'<a href="{url}" class="button">📤 Upload CSV/Excel</a>')

    upload_files_link.short_description = "Upload Files"
    download_links.allow_tags = True


admin.site.register(Company, CompanyAdmin)
admin.site.register(CompanyUpload)
